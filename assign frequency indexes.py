"""
   This file takes a workbook with a copy of two sheets from a workbook:
        - Top 4000 Edited
        - Top 4000 Original (frequency)

    This program will:
        1. assign an index to each word
        2. sort in alphabetical order
        3. Clean up the strings
        4. Produce the final Top 4000 most frequent words but with indexes for each word
        **Does not remove duplicates (to be done in excel)


    Indexes will be assigned in a categorical fashion:
        - 1-621 = 1
        - 622-1243 = 2
        - 1244-1862 = 3
        --> 1 is highest frequency, 3 is lowest frequency

    Pseudocode:

        load workbook
        define sheet name
        clean up main column
        (excel) remove duplicates
        assign indexes to each word
        compare to existing final product
            remove anything not in final product
"""

import unicodedata
from openpyxl import load_workbook

#removes parentheses
def cut_parentheses(string):
    if '(' in string:  # don't cut strings that are in good condition
        start = string.find('(')
        end   = string.find(')')
        string = string[0:start] + string[(end+2):]
    return string

#removes first space from a cell in the author's faulty spreadsheet
def removeFirstSpace(string):
    start = string.find(' ') + 1
    if start == 1:
        string = string[start:]
    return string

#makes a cell have only one word instead of multiple words/definitions
def makeOneWord(string):
    for char in string:
        if (char.isalpha()) == False:
            start = string.find(char)
            string = string[0:start]
            break  # to save time
    return string

#cleans column B
#receives a newly created excel spreadsheet
#returns nothing
def clean_column_B(sheet):
    # loop through every cell in column "B" only
    for col_cells in sheet.iter_cols(min_col=2, max_col=2):
        for cell in col_cells:
            cell.value = unicodedata.normalize("NFKD", cell.value)
            cell.value = cut_parentheses(cell.value)
            cell.value = removeFirstSpace(cell.value)
            cell.value = makeOneWord(cell.value)

#assign indexes to column C from column B value
#receives a newly created excel spreadsheet
#returns nothing
def assign_indexes(sheet):
    #loop through every cell in column "C" only

    #  1-1203 gets a 1
    #  1204 - 2406 gets a 2
    #  2407 - 3609 gets a 3
    for col_cells in sheet.iter_cols(min_col=3, max_col=3):
        i = 1
        for cell in col_cells:
            if i < 1204:
                cell.value = 1
            elif i > 1204 and i < 2407:
                cell.value = 2
            else:
                cell.value = 3
            i = i + 1


# load workbook
wb = load_workbook(filename="container.xlsx")
words_by_frequency_sheet = wb.get_sheet_by_name('Words by Frequency')
sheet_vals = tuple(words_by_frequency_sheet.columns)

#split sheet_vals into column cell object lists
column_A, column_B = sheet_vals[0], sheet_vals[1]

#clean column B
clean_column_B(words_by_frequency_sheet)

#assign indexes
assign_indexes(words_by_frequency_sheet)

#save workbook
wb.save('container.xlsx')
