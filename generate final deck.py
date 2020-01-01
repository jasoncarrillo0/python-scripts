from openpyxl import load_workbook
import genanki
import random
import unicodedata
"""
This program utilizes the "genanki" library created by a user of the flash card system. You can use this library
to create an Anki flashcard deck. This script takes in the Top 1500 most frequently used Lebanese Arabic words
from a cleaned excel spreadsheet, and exports an Anki deck to be used in the Anki desktop/mobile app.

pseudocode

1. be able to generate a random number that's really big without repitions
2. setup note type
3. setup the model
4. make a deck to fill
5. figure out note types
"""

beg = 1000000000
end = 9999999999
model_id = random.randrange(1 << 30, 1 << 31)
id_list = []
for i in range(0, 5000):
    randomNum = random.randint(beg,end)
    while randomNum in id_list:
        randomNum = random.randint(beg, end)
    id_list.append(randomNum)


model = genanki.Model(
  model_id,
  'Simple Model',
  fields=[
    {'name': 'English Word'},
    {'name': 'Arabic Spelling'},
    {'name': 'Phonetic Spelling'},
    {'name': 'Remarks'},
  ],
  templates=[
      {
        'name': 'Only Card',
          'qfmt': 'How do you say:<div style="font-family: Arial; font-size: 40px; padding: 20px;">{{Arabic Spelling}}</div>',
          'afmt': '{{FrontSide}}<hr id=answer><div style="font-family: Arial; font-size: 20px; padding: 20px;">{{Meaning}}</div><div style="font-family: Arial; font-size: 20px; padding: 20px;"><em>{{Phonetic Spelling}}</em></div><div style="font-family: Arial; font-size: 20px; padding: 20px;">{{Audio}}</div>',
      },
  ])

my_deck = genanki.Deck(
  id_list[-1],
  'Top 1500')
del id_list[-1]

wb = load_workbook("all_words.xlsx")
sheet = wb.get_sheet_by_name('ALL WORDS')
big_list = []
row_list = []

for row in sheet.iter_rows(min_row=2, max_col=4, max_row=1507):
    for cell in row:
        cell.value = unicodedata.normalize("NFKD", cell.value)
        row_list.append(cell.value)
    big_list.append(row_list)
    row_list = []

i = 0
for item in big_list:
    note = genanki.Note(
        model=model,
        fields=[item[0], item[1], item[2], item[3]]
    )
    my_deck.add_note(note)
    i = i + 1
i = 0

genanki.Package(my_deck).write_to_file("Top 1500.apkg")






