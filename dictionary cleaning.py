"""
Purpose of This Program:

Program is meant to:
    1. Clean up an excel spreadsheet that I created from an online Anki Flashcard deck of Arabic words
    2. Create consistency in english definition columns (e.g., every english word will have only one english word)
    3. Create consistency in arabic translation columns
    4. Transfer arabic words from author's sheet to custom made excel sheet

Conventions author uses for designating word type:
    - (m, fm, p) for describing a person
    - (s, pl)

Possible english cases to clean:
test1 = "colder or coldest jason loves (s, pl) i love you"
test2 = "how lucky (w/ pronoun me, you...)"
test3 = "class - in order, a line (s, pl)"
test4 = "dark (as in dark color)"
test5 = "game, toy, doll (s, pl)"
test6 = "cold or coldest"
test7 = "jason loves me too because I love him the same" # should return one word
test8 = "jason and layal also string doesn't have an or or and" #should return first three words
test9 = "jason (money)"
test10 = "ItWorkedHomie"
testList = [test1, test2, test3, test4, test5, test6, test7, test8, test9, test10]
"""

from openpyxl import load_workbook
import unicodedata
import re


# turns this: "game, toy, doll (s, pl)"
# into this:  "game (s,pl)"
# function returns a string
def cutImmediateComma(string):
    if "(" in string and ")" in string:
        openParenthesesIdx = string.find("(")
        closeParenthesesIdx = string.find(")")
        afterFirstWordIdx = 0
        for char in string:
            if char.isalpha() == False and char.isspace() == False:
                afterFirstWordIdx = string.find(char)
                break
        parentheses = "".join(string[openParenthesesIdx:(closeParenthesesIdx + 1)].split())
        return string[0:afterFirstWordIdx] + " " + parentheses

#a more thorough way to find "nth" occurrence in a list
#receives a list (haystack), string to find (needle), and nth occurrence "n"
def find_nth(haystack, needle, n):
    start = haystack.find(needle)
    while start >= 0 and n > 1:
        start = haystack.find(needle, start+len(needle))
        n -= 1
    return start


#Handles all test cases in pseudocode.txt
#receives the english translation of an Arabic word as a string
#returns the correctly formatted string
def returnCleanCell(string):
    andOrList = ['and', 'or']
    whitespaces = sum(1 for match in re.finditer('\s+', string))

    #for strings with parentheses
    if "(" in string and ")" in string:
        openParenthesesIdx = string.find("(")
        closeParenthesesIdx = string.find(")")
        parentheses = "".join(string[openParenthesesIdx:(closeParenthesesIdx + 1)].split())
        wordList = re.sub("[^\w]()", " ", string).split()
        # so far, we have a list of words from the string, and the parentheses with removed space
        if whitespaces > 2:
            secondSpaceIdx = find_nth(string, " ", 2)
            if wordList[1] in andOrList:  #handles "colder or coldest jason loves (s, pl) i love you"
                return wordList[0] + " " + wordList[1] + " " + wordList[2] + " " + parentheses
            elif wordList[1] not in andOrList and secondSpaceIdx == openParenthesesIdx:
                return wordList[0] + " " + wordList[1] + " " + parentheses
            else:                         # handles "colder (money)"
                return wordList[0] + " " + parentheses
        else:
            return wordList[0] + " " + parentheses

    #for strings without parentheses
    else:
        wordList = string.split()
        wordListLength = len(wordList)
        if whitespaces > 2 and wordListLength > 3 and wordList[1] in andOrList:
            # "cold or coldest jason loves me"
            return wordList[0] + " " + wordList[1] + " " + wordList[2]
        elif whitespaces == 2 and wordList[1] in andOrList and wordListLength == 3:  # "cold or coldest"
            return wordList[0] + " " + wordList[1] + " " + wordList[2]
        else:
            return wordList[0]
        return string


#removes hidden characters of excel cell values of the whole column
#replaces cell values with clean english definition of the whole column
#receives the excel sheet, and column number to iterate over
#returns nothing
def cleanColumn(sheet, columnNum):
    for col_cells in sheet.iter_cols(min_col=columnNum, max_col=columnNum):
        for cell in col_cells:
            cell.value = unicodedata.normalize("NFKD", cell.value)
            cell.value = returnCleanCell(cell.value)


# appends cleaned arabic values from author's sheet to custom sheet
# receives author's excel sheet, and custom excel sheet
# returns nothing
def appendArabicValues(ankiSheet, allWordsSheet):
    # create list from author's sheet column A
    ankiEnglishList = []
    for col_cells in ankiSheet.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in col_cells:
            ankiEnglishList.append(cell.value)
    ankiEnglishList = [sentence.split() for sentence in ankiEnglishList]
    # ankiEnglishList is now a list of lists [ [word, word], [word, word, word] ]

    for row in allWordsSheet.iter_rows(min_row=2, min_col=1, max_col=1):
        # variables
        rowNum = 2
        column_as_char_b = 'B'
        andOrList = ['and', 'or']

        for cell in row:
            # make list of words in english cell
            englishCell = cell.value.split(' ')

            # wordList is a "list" type, ankiEnglishList is a "list of lists"
            for wordList in ankiEnglishList:
                if englishCell[0] == wordList[0]:
                    arabicWordRowNum = ankiEnglishList.index(wordList) + 2
                    arabicWordCellRef = column_as_char_b + str(arabicWordRowNum)
                    englishCellRef = column_as_char_b + str(cell.row)
                    arabic_list = ankiSheet[arabicWordCellRef].value.split('،')
                    arabic_listLength = len(arabic_list)
                    arabic_list.reverse()
                    """
                    Debugging and Testing

                    print(
                        "searching for englishCell[0] which is: " + englishCell[0] + "\n"
                        "search for it in wordList: " + str(wordList) + "\n" +
                        "Found; it's in row: (arabicWordRowNum):  " + str(arabicWordRowNum) + "\n" +
                        "To reference, we made the cell reference: " + arabicWordCellRef + "\n" +
                        "list of arabic words from cell reference: " + str(arabic_list) + "\n" 
                        "we will now write the arabic word to:     " + str(englishCellRef) +"\n\n")
                    """
                    if not allWordsSheet[englishCellRef].value:
                        if len(wordList) < 2 and arabic_listLength:
                            allWordsSheet[englishCellRef] = arabic_list[0]
                        elif len(wordList) == 2 and arabic_listLength > 2:
                            allWordsSheet[englishCellRef] = arabic_list[0] + '/' + arabic_list[1]
                        elif len(wordList) == 2 and arabic_listLength == 2:
                            allWordsSheet[englishCellRef] = arabic_list[0] + '/' + arabic_list[1]
                        elif len(wordList) == 2 and arabic_listLength < 2:
                            allWordsSheet[englishCellRef] = arabic_list[0]
                        elif len(wordList) > 2 and wordList[1] in andOrList and len(arabic_list) > 1:
                            allWordsSheet[englishCellRef] = arabic_list[0] + '/' + arabic_list[1]
                        else:
                            allWordsSheet[englishCellRef] = arabic_list[0]
                        break
                rowNum = rowNum + 1
            rowNum = 2




# load downloaded workbook
wb_Anki   = load_workbook(filename="Anki Levantine Arabid Dictionary.xlsx")
ankiSheet = wb_Anki.get_sheet_by_name('English Front Deck')

# load custom sheet
wb_AllWords = load_workbook(filename="Arabic Master Copy.xlsx")
wordSheet   = wb_AllWords.active

# clean author's english definition column
cleanColumn(ankiSheet, 1)

#append arabic values
appendArabicValues(ankiSheet, wordSheet)

#save workbooks
wb_AllWords.save('/Users/jasoncarrillo/Desktop/ARABIC LEARNING/test_ALLWORDS.xlsx')
wb_Anki.save('/Users/jasoncarrillo/Desktop/ARABIC LEARNING/test_ANKIDECK.xlsx')
















